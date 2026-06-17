from flask import Flask, jsonify, request, abort
from flask_cors import CORS
from db import get_db, init_db
import os, io
from datetime import datetime

# Try importing xlsx support (optional)
try:
    import openpyxl
    XLSX_SUPPORT = True
except ImportError:
    XLSX_SUPPORT = False

app = Flask(__name__)
# CORS(app, origins=["http://localhost:5173", "http://localhost:4173", "http://127.0.0.1:5173"])
CORS(app, origins=[
    "http://localhost:5173",
    "http://localhost:4173",
    os.environ.get("FRONTEND_URL", "")
])
init_db()
# ── HELPERS ──────────────────────────────────────────────────────────────────

def row_to_dict(row):
    return dict(row) if row else None

def rows_to_list(rows):
    return [dict(r) for r in rows]

def ok(data=None, **kwargs):
    resp = {"ok": True}
    if data is not None:
        resp["data"] = data
    resp.update(kwargs)
    return jsonify(resp)

def err(msg, status=400):
    return jsonify({"ok": False, "error": msg}), status

# ── SCHEDULE / EVENTS ─────────────────────────────────────────────────────────

@app.get("/api/events")
def get_events():
    db = get_db()
    rows = db.execute("SELECT * FROM events ORDER BY day, time").fetchall()
    db.close()
    return ok(rows_to_list(rows))

@app.post("/api/events")
def create_event():
    d = request.json or {}
    name = (d.get("name") or "").strip()
    day  = d.get("day")
    time = (d.get("time") or "").strip()
    cat  = (d.get("cat") or "personal").strip()
    if not name or day is None or not time:
        return err("name, day and time are required")
    db = get_db()
    cur = db.execute("INSERT INTO events(name,day,time,cat) VALUES(?,?,?,?)", (name, day, time, cat))
    db.commit()
    row = db.execute("SELECT * FROM events WHERE id=?", (cur.lastrowid,)).fetchone()
    db.close()
    return ok(row_to_dict(row)), 201

@app.put("/api/events/<int:eid>")
def update_event(eid):
    d = request.json or {}
    db = get_db()
    existing = db.execute("SELECT * FROM events WHERE id=?", (eid,)).fetchone()
    if not existing:
        db.close(); return err("Not found", 404)
    name = d.get("name", existing["name"])
    day  = d.get("day",  existing["day"])
    time = d.get("time", existing["time"])
    cat  = d.get("cat",  existing["cat"])
    db.execute("UPDATE events SET name=?,day=?,time=?,cat=? WHERE id=?", (name,day,time,cat,eid))
    db.commit()
    row = db.execute("SELECT * FROM events WHERE id=?", (eid,)).fetchone()
    db.close()
    return ok(row_to_dict(row))

@app.delete("/api/events/<int:eid>")
def delete_event(eid):
    db = get_db()
    db.execute("DELETE FROM events WHERE id=?", (eid,))
    db.commit(); db.close()
    return ok()

# ── WORKOUTS ──────────────────────────────────────────────────────────────────

@app.get("/api/workouts")
def get_workouts():
    db = get_db()
    workouts = rows_to_list(db.execute("SELECT * FROM workouts ORDER BY CASE day WHEN 'Mon' THEN 0 WHEN 'Tue' THEN 1 WHEN 'Wed' THEN 2 WHEN 'Thu' THEN 3 WHEN 'Fri' THEN 4 WHEN 'Sat' THEN 5 ELSE 6 END").fetchall())
    exercises = rows_to_list(db.execute("SELECT * FROM exercises ORDER BY workout_id, sort_order").fetchall())
    db.close()
    ex_map = {}
    for ex in exercises:
        ex_map.setdefault(ex["workout_id"], []).append(ex)
    for w in workouts:
        w["exercises"] = ex_map.get(w["id"], [])
    return ok(workouts)

@app.post("/api/workouts")
def create_workout():
    d = request.json or {}
    day  = (d.get("day") or "").strip()
    wtype = (d.get("type") or "rest").strip()
    exs  = d.get("exercises", [])
    if not day:
        return err("day is required")
    db = get_db()
    # Replace existing workout for that day
    existing = db.execute("SELECT id FROM workouts WHERE day=?", (day,)).fetchone()
    if existing:
        db.execute("DELETE FROM workouts WHERE id=?", (existing["id"],))
    cur = db.execute("INSERT INTO workouts(day,type) VALUES(?,?)", (day, wtype))
    wid = cur.lastrowid
    for i, ex in enumerate(exs):
        db.execute("INSERT INTO exercises(workout_id,name,sets,weight,sort_order) VALUES(?,?,?,?,?)",
                   (wid, ex.get("name",""), ex.get("sets",""), ex.get("weight",""), i))
    db.commit()
    row = db.execute("SELECT * FROM workouts WHERE id=?", (wid,)).fetchone()
    result = row_to_dict(row)
    result["exercises"] = rows_to_list(db.execute("SELECT * FROM exercises WHERE workout_id=? ORDER BY sort_order", (wid,)).fetchall())
    db.close()
    return ok(result), 201

@app.post("/api/workouts/<int:wid>/exercises")
def add_exercise(wid):
    d = request.json or {}
    name = (d.get("name") or "").strip()
    sets = (d.get("sets") or "").strip()
    weight = (d.get("weight") or "").strip()
    if not name:
        return err("name is required")
    db = get_db()
    count = db.execute("SELECT COUNT(*) FROM exercises WHERE workout_id=?", (wid,)).fetchone()[0]
    db.execute("INSERT INTO exercises(workout_id,name,sets,weight,sort_order) VALUES(?,?,?,?,?)",
               (wid, name, sets, weight, count))
    db.commit()
    rows = rows_to_list(db.execute("SELECT * FROM exercises WHERE workout_id=? ORDER BY sort_order", (wid,)).fetchall())
    db.close()
    return ok(rows), 201

@app.delete("/api/workouts/<int:wid>")
def delete_workout(wid):
    db = get_db()
    db.execute("DELETE FROM workouts WHERE id=?", (wid,))
    db.commit(); db.close()
    return ok()

# ── TRANSACTIONS ──────────────────────────────────────────────────────────────

@app.get("/api/transactions")
def get_transactions():
    db = get_db()
    rows = db.execute("SELECT * FROM transactions ORDER BY created_at DESC").fetchall()
    db.close()
    return ok(rows_to_list(rows))

@app.post("/api/transactions")
def create_transaction():
    d = request.json or {}
    name   = (d.get("name") or "").strip()
    amount = d.get("amount")
    cat    = (d.get("cat") or "Other").strip()
    date   = (d.get("date") or datetime.now().strftime("%b %d")).strip()
    if not name or amount is None:
        return err("name and amount are required")
    try:
        amount = float(amount)
    except (TypeError, ValueError):
        return err("amount must be a number")
    db = get_db()
    cur = db.execute("INSERT INTO transactions(name,amount,cat,date) VALUES(?,?,?,?)", (name, amount, cat, date))
    db.commit()
    row = db.execute("SELECT * FROM transactions WHERE id=?", (cur.lastrowid,)).fetchone()
    db.close()
    return ok(row_to_dict(row)), 201

@app.put("/api/transactions/<int:tid>")
def update_transaction(tid):
    d = request.json or {}
    db = get_db()
    existing = db.execute("SELECT * FROM transactions WHERE id=?", (tid,)).fetchone()
    if not existing:
        db.close(); return err("Not found", 404)
    name   = d.get("name",   existing["name"])
    amount = d.get("amount", existing["amount"])
    cat    = d.get("cat",    existing["cat"])
    date   = d.get("date",   existing["date"])
    db.execute("UPDATE transactions SET name=?,amount=?,cat=?,date=? WHERE id=?", (name,amount,cat,date,tid))
    db.commit()
    row = db.execute("SELECT * FROM transactions WHERE id=?", (tid,)).fetchone()
    db.close()
    return ok(row_to_dict(row))

@app.delete("/api/transactions/<int:tid>")
def delete_transaction(tid):
    db = get_db()
    db.execute("DELETE FROM transactions WHERE id=?", (tid,))
    db.commit(); db.close()
    return ok()

# ── GROCERIES ─────────────────────────────────────────────────────────────────

@app.get("/api/groceries")
def get_groceries():
    db = get_db()
    rows = db.execute("SELECT * FROM groceries ORDER BY sort_order, created_at").fetchall()
    db.close()
    return ok(rows_to_list(rows))

@app.post("/api/groceries")
def create_grocery():
    d = request.json or {}
    name  = (d.get("name") or "").strip()
    cat   = (d.get("cat") or "Other").strip()
    price = float(d.get("price") or 0)
    if not name:
        return err("name is required")
    db = get_db()
    count = db.execute("SELECT COUNT(*) FROM groceries").fetchone()[0]
    cur = db.execute("INSERT INTO groceries(name,cat,price,done,sort_order) VALUES(?,?,?,0,?)", (name, cat, price, count))
    db.commit()
    row = db.execute("SELECT * FROM groceries WHERE id=?", (cur.lastrowid,)).fetchone()
    db.close()
    return ok(row_to_dict(row)), 201

@app.patch("/api/groceries/<int:gid>")
def patch_grocery(gid):
    d = request.json or {}
    db = get_db()
    existing = db.execute("SELECT * FROM groceries WHERE id=?", (gid,)).fetchone()
    if not existing:
        db.close(); return err("Not found", 404)
    name  = d.get("name",  existing["name"])
    cat   = d.get("cat",   existing["cat"])
    price = d.get("price", existing["price"])
    done  = int(d.get("done", existing["done"]))
    db.execute("UPDATE groceries SET name=?,cat=?,price=?,done=? WHERE id=?", (name,cat,price,done,gid))
    db.commit()
    row = db.execute("SELECT * FROM groceries WHERE id=?", (gid,)).fetchone()
    db.close()
    return ok(row_to_dict(row))

@app.delete("/api/groceries/<int:gid>")
def delete_grocery(gid):
    db = get_db()
    db.execute("DELETE FROM groceries WHERE id=?", (gid,))
    db.commit(); db.close()
    return ok()

# ── HABITS ────────────────────────────────────────────────────────────────────

@app.get("/api/habits")
def get_habits():
    db = get_db()
    habits = rows_to_list(db.execute("SELECT * FROM habits ORDER BY sort_order, created_at").fetchall())
    logs   = rows_to_list(db.execute("SELECT * FROM habit_logs").fetchall())
    db.close()
    log_map = {}
    for l in logs:
        log_map.setdefault(l["habit_id"], []).append(l["log_date"])
    for h in habits:
        h["log"] = log_map.get(h["id"], [])
    return ok(habits)

@app.post("/api/habits")
def create_habit():
    d = request.json or {}
    name = (d.get("name") or "").strip()
    icon = (d.get("icon") or "✓").strip()
    if not name:
        return err("name is required")
    db = get_db()
    count = db.execute("SELECT COUNT(*) FROM habits").fetchone()[0]
    cur = db.execute("INSERT INTO habits(name,icon,sort_order) VALUES(?,?,?)", (name, icon, count))
    db.commit()
    row = dict(db.execute("SELECT * FROM habits WHERE id=?", (cur.lastrowid,)).fetchone())
    row["log"] = []
    db.close()
    return ok(row), 201

@app.delete("/api/habits/<int:hid>")
def delete_habit(hid):
    db = get_db()
    db.execute("DELETE FROM habits WHERE id=?", (hid,))
    db.commit(); db.close()
    return ok()

@app.post("/api/habits/<int:hid>/log")
def toggle_habit_log(hid):
    d = request.json or {}
    date = (d.get("date") or "").strip()
    if not date:
        return err("date is required")
    db = get_db()
    existing = db.execute("SELECT id FROM habit_logs WHERE habit_id=? AND log_date=?", (hid, date)).fetchone()
    if existing:
        db.execute("DELETE FROM habit_logs WHERE habit_id=? AND log_date=?", (hid, date))
        state = False
    else:
        db.execute("INSERT INTO habit_logs(habit_id,log_date) VALUES(?,?)", (hid, date))
        state = True
    db.commit()
    logs = [r["log_date"] for r in db.execute("SELECT log_date FROM habit_logs WHERE habit_id=? ORDER BY log_date", (hid,)).fetchall()]
    db.close()
    return ok({"logged": state, "log": logs})

# ── SETTINGS / BUDGET ─────────────────────────────────────────────────────────

@app.get("/api/settings")
def get_settings():
    db = get_db()
    rows = db.execute("SELECT * FROM settings").fetchall()
    db.close()
    return ok({r["key"]: r["value"] for r in rows})

@app.put("/api/settings/<key>")
def update_setting(key):
    d = request.json or {}
    value = str(d.get("value", ""))
    db = get_db()
    db.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", (key, value))
    db.commit(); db.close()
    return ok({"key": key, "value": value})

# ── EXCEL IMPORT ──────────────────────────────────────────────────────────────

@app.post("/api/import/excel")
def import_excel():
    if not XLSX_SUPPORT:
        return err("openpyxl not installed on server", 501)
    if "file" not in request.files:
        return err("No file uploaded")
    f = request.files["file"]
    if not f.filename.endswith((".xlsx", ".xls")):
        return err("File must be .xlsx or .xls")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        return err(f"Could not parse Excel file: {e}")

    db = get_db()
    imported = {}

    # Sheet: Transactions  (columns: name | amount | category | date)
    if "Transactions" in wb.sheetnames:
        ws = wb["Transactions"]
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, amount, cat, date = (row + (None,None,None,None))[:4]
            if not name or amount is None:
                continue
            try:
                amount = float(amount)
            except:
                continue
            cat  = str(cat or "Other").strip()
            date = str(date or datetime.now().strftime("%b %d")).strip()
            db.execute("INSERT INTO transactions(name,amount,cat,date) VALUES(?,?,?,?)", (str(name).strip(), amount, cat, date))
            count += 1
        imported["transactions"] = count

    # Sheet: Groceries  (columns: name | category | price | done)
    if "Groceries" in wb.sheetnames:
        ws = wb["Groceries"]
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, cat, price, done = (row + (None,None,None,None))[:4]
            if not name:
                continue
            cat   = str(cat or "Other").strip()
            price = float(price or 0)
            done  = 1 if str(done or "").lower() in ("1","yes","true","x") else 0
            sort  = db.execute("SELECT COUNT(*) FROM groceries").fetchone()[0]
            db.execute("INSERT INTO groceries(name,cat,price,done,sort_order) VALUES(?,?,?,?,?)", (str(name).strip(), cat, price, done, sort))
            count += 1
        imported["groceries"] = count

    # Sheet: Habits  (columns: name | icon)
    if "Habits" in wb.sheetnames:
        ws = wb["Habits"]
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, icon = (row + (None,None))[:2]
            if not name:
                continue
            icon  = str(icon or "✓").strip()
            sort  = db.execute("SELECT COUNT(*) FROM habits").fetchone()[0]
            db.execute("INSERT INTO habits(name,icon,sort_order) VALUES(?,?,?)", (str(name).strip(), icon, sort))
            count += 1
        imported["habits"] = count

    # Sheet: Schedule  (columns: name | day_index | time | category)
    if "Schedule" in wb.sheetnames:
        ws = wb["Schedule"]
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, day, time, cat = (row + (None,None,None,None))[:4]
            if not name or day is None:
                continue
            try:
                day = int(day)
            except:
                continue
            time = str(time or "09:00").strip()
            cat  = str(cat or "personal").strip()
            db.execute("INSERT INTO events(name,day,time,cat) VALUES(?,?,?,?)", (str(name).strip(), day, time, cat))
            count += 1
        imported["events"] = count

    db.commit(); db.close()
    return ok(imported)

# ── EXCEL EXPORT ──────────────────────────────────────────────────────────────

@app.get("/api/export/excel")
def export_excel():
    if not XLSX_SUPPORT:
        return err("openpyxl not installed", 501)
    from flask import send_file
    db = get_db()
    wb = openpyxl.Workbook()

    def add_sheet(name, headers, rows):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for row in rows:
            ws.append(row)

    add_sheet("Transactions",
        ["name","amount","category","date"],
        [(r["name"], r["amount"], r["cat"], r["date"])
         for r in db.execute("SELECT * FROM transactions ORDER BY created_at DESC").fetchall()])

    add_sheet("Groceries",
        ["name","category","price","done"],
        [(r["name"], r["cat"], r["price"], bool(r["done"]))
         for r in db.execute("SELECT * FROM groceries ORDER BY sort_order").fetchall()])

    add_sheet("Habits",
        ["name","icon"],
        [(r["name"], r["icon"])
         for r in db.execute("SELECT * FROM habits ORDER BY sort_order").fetchall()])

    add_sheet("Schedule",
        ["name","day_index","time","category"],
        [(r["name"], r["day"], r["time"], r["cat"])
         for r in db.execute("SELECT * FROM events ORDER BY day, time").fetchall()])

    del wb["Sheet"]  # remove default empty sheet
    db.close()

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     download_name="apex_data.xlsx", as_attachment=True)

# ── HEALTH ────────────────────────────────────────────────────────────────────

@app.get("/api/health")
def health():
    return ok({"status": "ok", "xlsx_support": XLSX_SUPPORT})

# ── MAIN ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", 8000))
    print(f"[APEX API] Running on http://localhost:{port}")
    app.run(host="0.0.0.0", port=port, debug=True)